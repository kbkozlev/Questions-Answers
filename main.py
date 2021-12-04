from openpyxl import load_workbook
import random
import os


def cls():
    os.system('cls' if os.name == 'nt' else 'clear')


wb = load_workbook("Questions&Answers.xlsx")  # Work Book
ws = wb['Sheet1']  # Work Sheet
questions = ws['A']  # Column
answers = ws['B']
questions_list = [questions[x].value for x in range(len(questions))]
answers_list = [answers[x].value for x in range(len(answers))]

rn = list(zip(questions_list, answers_list))
random.shuffle(rn)
questions_list, answers_list = zip(*rn)


def calc():
    number = 1
    calc.points = 0
    for x in range(len(questions_list)):
        input((str(number) + '. ' + questions_list[x]))
        number += 1
        p = input('\n' + '- ' + answers_list[x] + "\n" + '''
========================
  Is this correct? Y/N ''')
        if p == 'y':
            calc.points += 1
            cls()
        else:
            cls()
            continue


calc()

x = True

while x:
    result = round((calc.points / int(len(questions)) * 100), 1)
    print("You got " + str(calc.points) + ' | ' + str(len(questions)) + " correct.")
    print('This is ' + str(result) + '%\n')
    cn = input("Continue?...y/n ")
    cn.lower()
    if cn == 'y':
        random.shuffle(rn)
        questions_list, answers_list = zip(*rn)
        calc()
    else:
        x = False
        break
